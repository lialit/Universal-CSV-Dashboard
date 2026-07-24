from __future__ import annotations

from copy import copy
from io import BytesIO
import json
from typing import Mapping

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd

from app_core.insights import build_business_insights
from app_core.metrics import summarize_metric
from app_core.quality import (
    calculate_quality_score,
    quality_checks_table,
)
from app_core.report_themes import ReportTheme, get_report_theme


EXCEL_MAX_DATA_ROWS = 1_048_575
EXCEL_MAX_COLUMNS = 16_384


def validate_export_dimensions(rows: int, columns: int) -> None:
    if rows > EXCEL_MAX_DATA_ROWS:
        raise ValueError(
            f"Excel supports at most {EXCEL_MAX_DATA_ROWS:,} data rows "
            "per worksheet."
        )
    if columns > EXCEL_MAX_COLUMNS:
        raise ValueError(
            f"Excel supports at most {EXCEL_MAX_COLUMNS:,} columns "
            "per worksheet."
        )


def _safe_value(value):
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        value = value.item()
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _write_rows(worksheet, rows) -> None:
    for row in rows:
        worksheet.append([_safe_value(value) for value in row])


def _style_title(
    worksheet,
    title: str,
    theme: ReportTheme,
    last_column: int = 6,
) -> None:
    worksheet.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_column,
    )
    cell = worksheet.cell(1, 1, title)
    cell.fill = PatternFill(
        "solid",
        fgColor=theme.title_background,
    )
    cell.font = Font(
        name="Aptos Display",
        size=20,
        bold=True,
        color=theme.title_text,
    )
    cell.alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 34


def _style_header(
    worksheet,
    row: int,
    columns: int,
    theme: ReportTheme,
) -> None:
    for cell in worksheet[row][:columns]:
        cell.fill = PatternFill(
            "solid",
            fgColor=theme.accent,
        )
        cell.font = Font(
            name="Aptos",
            bold=True,
            color=theme.accent_text,
        )
        cell.alignment = Alignment(
            vertical="center",
            wrap_text=True,
        )
    worksheet.row_dimensions[row].height = 24


def _finish_sheet(
    worksheet,
    theme: ReportTheme,
    widths: Mapping[str, float] | None = None,
) -> None:
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_properties.tabColor = theme.accent
    worksheet.freeze_panes = "A3"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0

    if widths:
        for column, width in widths.items():
            worksheet.column_dimensions[column].width = width

    surface_fill = PatternFill(
        "solid",
        fgColor=theme.surface,
    )
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=worksheet.max_column,
    ):
        for cell in row:
            if cell.fill.fill_type is None:
                cell.fill = surface_fill
            if cell.font.color is None or cell.font.color.type == "theme":
                font = copy(cell.font)
                font.color = theme.text
                cell.font = font


def _add_table(
    worksheet,
    reference: str,
    name: str,
    theme: ReportTheme,
) -> None:
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name=theme.excel_table_style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _build_overview(
    workbook: Workbook,
    dataframe: pd.DataFrame,
    config: Mapping[str, object],
    source_name: str,
    theme: ReportTheme,
) -> None:
    worksheet = workbook.active
    worksheet.title = "Overview"
    _style_title(
        worksheet,
        "Universal CSV Dashboard — Analysis Export",
        theme,
    )

    metric = str(config["metric_column"])
    metric_summary = summarize_metric(dataframe, metric)
    metric_column_number = dataframe.columns.get_loc(metric) + 1
    metric_letter = get_column_letter(metric_column_number)
    data_end_row = len(dataframe) + 1
    metric_range = (
        f"'Data'!{metric_letter}2:{metric_letter}{data_end_row}"
    )

    worksheet.append([])
    worksheet.append(["Source file", _safe_value(source_name)])
    worksheet.append(["Rows", len(dataframe)])
    worksheet.append(["Columns", len(dataframe.columns)])
    worksheet.append(["Primary metric", _safe_value(metric)])
    worksheet.append(["Aggregation", config.get("aggregation", "Sum")])

    worksheet.append([])
    worksheet.append(["KPI", "Value", "Source / formula"])
    _style_header(worksheet, 9, 3, theme)
    kpi_rows = (
        ("Total", f"=SUM({metric_range})", metric_range),
        ("Average", f"=AVERAGE({metric_range})", metric_range),
        ("Median", f"=MEDIAN({metric_range})", metric_range),
        ("Minimum", f"=MIN({metric_range})", metric_range),
        ("Maximum", f"=MAX({metric_range})", metric_range),
        ("Non-null count", f"=COUNT({metric_range})", metric_range),
    )
    for label, formula, source in kpi_rows:
        worksheet.append([label, formula, source])

    worksheet["E3"] = "Calculation check"
    worksheet["F3"] = metric_summary.total
    worksheet["E4"] = "Generated locally"
    worksheet["F4"] = "Yes"
    worksheet["E5"] = "Report theme"
    worksheet["F5"] = theme.name
    for cell in ("E3", "E4", "E5"):
        worksheet[cell].font = Font(
            bold=True,
            color=theme.text,
        )
    worksheet["F3"].number_format = "#,##0.00"

    thin = Side(style="thin", color="D9E2F1")
    for row in worksheet.iter_rows(
        min_row=3,
        max_row=15,
        min_col=1,
        max_col=6,
    ):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.font = Font(
                name="Aptos",
                color=theme.text,
                bold=cell.font.bold,
            )
    _style_header(worksheet, 9, 3, theme)
    for row in range(10, 16):
        worksheet.cell(row, 2).number_format = "#,##0.00"

    _finish_sheet(
        worksheet,
        theme,
        {
            "A": 24,
            "B": 20,
            "C": 34,
            "D": 4,
            "E": 22,
            "F": 20,
        },
    )


def _build_quality(
    workbook: Workbook,
    dataframe: pd.DataFrame,
    theme: ReportTheme,
) -> None:
    report = calculate_quality_score(dataframe)
    worksheet = workbook.create_sheet("Data Quality")
    _style_title(worksheet, "Data Quality", theme, 7)
    worksheet.append([])
    worksheet.append(["Score", report.score])
    worksheet.append(["Status", report.status])
    worksheet.append(["Detected issues", report.issue_count])
    worksheet.append([])

    table = quality_checks_table(report)
    worksheet.append(list(table.columns))
    _style_header(worksheet, 7, len(table.columns), theme)
    _write_rows(worksheet, table.itertuples(index=False, name=None))

    _add_table(
        worksheet,
        f"A7:G{worksheet.max_row}",
        "DataQualityChecks",
        theme,
    )
    worksheet["B3"].number_format = "0.0"
    worksheet["B4"].fill = PatternFill(
        "solid",
        fgColor=(
            theme.success_soft
            if report.score >= 90
            else theme.warning_soft
        ),
    )
    _finish_sheet(
        worksheet,
        theme,
        {
            "A": 24,
            "B": 14,
            "C": 12,
            "D": 18,
            "E": 12,
            "F": 46,
            "G": 48,
        },
    )


def _build_insights(
    workbook: Workbook,
    dataframe: pd.DataFrame,
    config: Mapping[str, object],
    theme: ReportTheme,
) -> None:
    report = build_business_insights(dataframe, config)
    worksheet = workbook.create_sheet("Business Insights")
    _style_title(worksheet, "Business Insights", theme, 8)
    worksheet.append([])
    worksheet.append(
        [
            "Type",
            "Title",
            "Observation",
            "Interpretation",
            "Evidence",
            "Confidence",
            "Confidence basis",
            "Limitation",
        ]
    )
    _style_header(worksheet, 3, 8, theme)

    if report.insights:
        _write_rows(
            worksheet,
            (
                (
                    insight.insight_type,
                    insight.title,
                    insight.observation,
                    insight.interpretation,
                    insight.evidence,
                    insight.confidence,
                    insight.confidence_reason,
                    insight.limitation,
                )
                for insight in report.insights
            ),
        )
        _add_table(
            worksheet,
            f"A3:H{worksheet.max_row}",
            "BusinessInsights",
            theme,
        )
    else:
        worksheet.append(
            [
                "Information",
                "No material insights",
                "No rule crossed the current thresholds.",
                None,
                None,
                None,
                None,
                None,
            ]
        )

    for row in worksheet.iter_rows(
        min_row=4,
        max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
        worksheet.row_dimensions[row[0].row].height = 72

    _finish_sheet(
        worksheet,
        theme,
        {
            "A": 16,
            "B": 28,
            "C": 42,
            "D": 42,
            "E": 46,
            "F": 14,
            "G": 45,
            "H": 48,
        },
    )


def _build_configuration(
    workbook: Workbook,
    config: Mapping[str, object],
    theme: ReportTheme,
) -> None:
    worksheet = workbook.create_sheet("Configuration")
    _style_title(
        worksheet,
        "Dashboard Configuration",
        theme,
        3,
    )
    worksheet.append([])
    worksheet.append(["Setting", "Value", "Machine-readable value"])
    _style_header(worksheet, 3, 3, theme)

    for key, value in config.items():
        display_value = (
            ", ".join(str(item) for item in value)
            if isinstance(value, list)
            else value
        )
        _write_rows(
            worksheet,
            [
                (
                    key,
                    display_value,
                    json.dumps(value, ensure_ascii=False),
                )
            ],
        )

    _add_table(
        worksheet,
        f"A3:C{worksheet.max_row}",
        "DashboardConfiguration",
        theme,
    )
    _finish_sheet(
        worksheet,
        theme,
        {"A": 24, "B": 42, "C": 58},
    )


def _build_data(
    workbook: Workbook,
    dataframe: pd.DataFrame,
    theme: ReportTheme,
) -> None:
    worksheet = workbook.create_sheet("Data")
    worksheet.append(
        [_safe_value(str(column)) for column in dataframe.columns]
    )
    _write_rows(
        worksheet,
        dataframe.itertuples(index=False, name=None),
    )
    _style_header(
        worksheet,
        1,
        len(dataframe.columns),
        theme,
    )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    for index, column in enumerate(dataframe.columns, start=1):
        letter = get_column_letter(index)
        width = min(max(len(str(column)) + 2, 12), 28)
        worksheet.column_dimensions[letter].width = width
        if pd.api.types.is_datetime64_any_dtype(dataframe[column]):
            for cell in worksheet[letter][1:]:
                cell.number_format = "yyyy-mm-dd hh:mm"

    _add_table(
        worksheet,
        f"A1:{get_column_letter(len(dataframe.columns))}"
        f"{len(dataframe) + 1}",
        "ExportedData",
        theme,
    )
    _finish_sheet(worksheet, theme)
    worksheet.freeze_panes = "A2"


def _build_methodology(
    workbook: Workbook,
    theme: ReportTheme,
) -> None:
    worksheet = workbook.create_sheet("Methodology")
    _style_title(
        worksheet,
        "Methodology and Responsible Use",
        theme,
        3,
    )
    worksheet.append([])
    worksheet.append(["Area", "Method", "Important limitation"])
    _style_header(worksheet, 3, 3, theme)
    rows = (
        (
            "Data Quality",
            "Completeness 50%, duplicate-free rows 30%, type validity 20%.",
            "Technical quality does not prove business accuracy.",
        ),
        (
            "Trend",
            "Compares average daily metric in the first and second half.",
            "The comparison does not establish seasonality or causation.",
        ),
        (
            "Anomaly",
            "Flags values outside 1.5× the interquartile range.",
            "A flagged value is not automatically an error.",
        ),
        (
            "Relationship",
            "Reports Pearson correlation when absolute correlation ≥ 0.50.",
            "Correlation does not prove causation.",
        ),
        (
            "Confidence",
            "Uses sample size, metric completeness and Data Quality Score.",
            "The label describes evidence reliability, not certainty.",
        ),
        (
            "Privacy",
            "Workbook is generated locally in the Streamlit session.",
            "Recipients of the workbook can access all exported data.",
        ),
    )
    _write_rows(worksheet, rows)
    _add_table(
        worksheet,
        f"A3:C{worksheet.max_row}",
        "ExportMethodology",
        theme,
    )
    for row in worksheet.iter_rows(
        min_row=4,
        max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
        worksheet.row_dimensions[row[0].row].height = 48
    _finish_sheet(
        worksheet,
        theme,
        {"A": 20, "B": 58, "C": 58},
    )


def build_excel_report(
    dataframe: pd.DataFrame,
    config: Mapping[str, object],
    source_name: str = "uploaded.csv",
    theme_name: str = "Corporate",
) -> bytes:
    """Build a styled, traceable Excel analysis workbook."""

    validate_export_dimensions(
        rows=len(dataframe),
        columns=len(dataframe.columns),
    )

    metric = config.get("metric_column")
    if not isinstance(metric, str) or metric not in dataframe.columns:
        raise ValueError(
            "A valid primary metric is required for Excel export."
        )

    theme = get_report_theme(theme_name)
    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    _build_overview(
        workbook,
        dataframe,
        config,
        source_name,
        theme,
    )
    _build_quality(workbook, dataframe, theme)
    _build_insights(workbook, dataframe, config, theme)
    _build_configuration(workbook, config, theme)
    _build_data(workbook, dataframe, theme)
    _build_methodology(workbook, theme)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
