from io import BytesIO

from openpyxl import load_workbook
import pandas as pd
from pypdf import PdfReader

from app_core.export_naming import export_filename, safe_source_label
from app_core.exports import build_excel_report
from app_core.pdf_exports import build_pdf_report


def config_for(metric: str = "sales") -> dict:
    return {
        "date_column": "date",
        "category_column": "category",
        "metric_column": metric,
        "numeric_columns": [metric],
        "aggregation": "Sum",
    }


def test_export_filename_is_deterministic_and_filesystem_safe() -> None:
    source = 'C:\\private\\Quarter 1: Sales?.csv'

    assert safe_source_label(source) == "Quarter 1: Sales?.csv"
    assert export_filename(source, "analysis", "xlsx") == (
        "Quarter_1_Sales_analysis.xlsx"
    )
    assert export_filename(source, "executive_report", ".pdf") == (
        "Quarter_1_Sales_executive_report.pdf"
    )


def test_excel_handles_missing_values_long_labels_and_wide_data() -> None:
    long_label = "Very long category label " * 8
    data = {
        "date": pd.date_range("2026-01-01", periods=8),
        "category": [long_label, "Short"] * 4,
        "sales": [100.0, None, 120.0, 130.0, None, 150.0, 160.0, 170.0],
    }
    for index in range(30):
        data[f"extra_metric_{index:02d}"] = [index + row for row in range(8)]
    dataframe = pd.DataFrame(data)

    content = build_excel_report(
        dataframe,
        config_for(),
        "wide_missing.csv",
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    sheet = workbook["Data"]

    assert sheet.max_column == len(dataframe.columns)
    assert sheet.max_row == len(dataframe) + 1
    assert sheet["C3"].value is None
    assert sheet["B2"].value == long_label
    assert sheet.freeze_panes == "A2"


def test_pdf_handles_long_labels_missing_values_and_safe_source_label() -> None:
    dataframe = pd.DataFrame(
        {
            "date": pd.date_range("2026-01-01", periods=40),
            "category": ["Region with a deliberately long business label"] * 40,
            "sales": [float(index + 1) if index % 7 else None for index in range(40)],
        }
    )
    source = safe_source_label("/Users/private/customer/report.csv")
    content = build_pdf_report(dataframe, config_for(), source)
    reader = PdfReader(BytesIO(content))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)

    assert content.startswith(b"%PDF-")
    assert len(reader.pages) >= 2
    assert "report.csv" in text
    assert "/Users/private/customer" not in text
