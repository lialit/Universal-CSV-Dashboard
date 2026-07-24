from io import BytesIO

from openpyxl import load_workbook
import pandas as pd
import pytest

from app_core.exports import (
    EXCEL_MAX_DATA_ROWS,
    build_excel_report,
    validate_export_dimensions,
)
from app_core.report_themes import REPORT_THEMES


def sample_dataframe() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "date": pd.date_range("2026-01-01", periods=10),
            "region": ["North", "South"] * 5,
            "sales": [10.0, 12.0, 9.0, 14.0, 11.0] * 2,
            "orders": list(range(1, 11)),
        }
    )


def sample_config() -> dict:
    return {
        "date_column": "date",
        "category_column": "region",
        "metric_column": "sales",
        "numeric_columns": ["sales", "orders"],
        "aggregation": "Sum",
        "kpi_cards": ["Total", "Average", "Median"],
        "chart_types": ["Time series", "Distribution"],
    }


def workbook(theme_name="Corporate"):
    content = build_excel_report(
        sample_dataframe(),
        sample_config(),
        "sample.csv",
        theme_name,
    )
    return load_workbook(BytesIO(content), data_only=False)


def test_export_contains_expected_worksheets():
    result = workbook()

    assert result.sheetnames == [
        "Overview",
        "Data Quality",
        "Business Insights",
        "Configuration",
        "Data",
        "Methodology",
    ]


def test_overview_kpis_are_traceable_formulas():
    result = workbook()
    overview = result["Overview"]

    assert overview["B10"].value.startswith("=SUM('Data'!")
    assert overview["B11"].value.startswith("=AVERAGE('Data'!")
    assert overview["B12"].value.startswith("=MEDIAN('Data'!")


def test_data_sheet_preserves_shape_and_dates():
    result = workbook()
    data = result["Data"]

    assert data.max_row == len(sample_dataframe()) + 1
    assert data.max_column == len(sample_dataframe().columns)
    assert data["A2"].is_date


def test_quality_and_methodology_are_included():
    result = workbook()

    assert result["Data Quality"]["A3"].value == "Score"
    assert result["Methodology"]["A4"].value == "Data Quality"
    assert "does not prove causation" in result["Methodology"]["C7"].value


def test_configuration_is_exported():
    result = workbook()
    configuration = result["Configuration"]
    keys = [
        configuration.cell(row, 1).value
        for row in range(4, configuration.max_row + 1)
    ]

    assert "metric_column" in keys
    assert "kpi_cards" in keys
    assert "chart_types" in keys


def test_formula_like_source_text_is_escaped():
    dataframe = sample_dataframe()
    dataframe.loc[0, "region"] = "=HYPERLINK(\"bad\")"
    content = build_excel_report(
        dataframe,
        sample_config(),
        "sample.csv",
    )
    result = load_workbook(BytesIO(content), data_only=False)

    assert result["Data"]["B2"].value.startswith("'=")
    assert result["Data"]["B2"].data_type == "s"


def test_excel_row_limit_fails_with_readable_error():
    with pytest.raises(ValueError, match="data rows"):
        validate_export_dimensions(
            EXCEL_MAX_DATA_ROWS + 1,
            4,
        )


def test_all_report_themes_are_applied_to_workbook():
    for theme_name, theme in REPORT_THEMES.items():
        result = workbook(theme_name)
        overview = result["Overview"]
        data = result["Data"]

        assert overview["F5"].value == theme_name
        assert overview["A1"].fill.fgColor.rgb.endswith(
            theme.title_background
        )
        assert overview["A1"].font.color.rgb.endswith(
            theme.title_text
        )
        assert data["A2"].fill.fgColor.rgb.endswith(
            theme.surface
        )
